import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("Date",           "date",         20),
    ("Source",         "source",       14),
    ("Nom",            "nom",          18),
    ("Prénom",         "prenom",       18),
    ("Email",          "email",        28),
    ("Téléphone",      "telephone",    16),
    ("Code postal",    "code_postal",  14),
    ("Assurance",      "assurance",    22),
    ("Message",        "message",      40),
    ("ID",             "id",           28),
]

NAV   = "002B5C"
RED   = "E8192C"
WHITE = "FFFFFF"
LIGHT = "F0F4F8"
BORDER = "CBD5E1"


def export_to_excel(leads: list[dict], cfg: dict = None) -> str:
    cfg = cfg or {}
    project = cfg.get("project_name", "Bot Universel")
    filename = cfg.get("export_filename", "leads")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = f"/tmp/{filename}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    tc = ws["A1"]
    tc.value = f"{project} — Leads export {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    tc.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    tc.fill      = PatternFill("solid", fgColor=NAV)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    hfill  = PatternFill("solid", fgColor=RED)
    hfont  = Font(bold=True, size=10, color=WHITE, name="Calibri")
    hborder = Border(bottom=Side(style="thin", color=BORDER), right=Side(style="thin", color=BORDER))

    for ci, (header, _, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=ci, value=header)
        c.font = hfont; c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = hborder
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 22

    # Data
    tb = Border(bottom=Side(style="thin", color=BORDER), right=Side(style="thin", color=BORDER))
    df = Font(size=10, name="Calibri")

    for ri, lead in enumerate(leads, 3):
        fill = PatternFill("solid", fgColor=LIGHT if ri % 2 == 0 else WHITE)
        for ci, (_, key, _) in enumerate(COLUMNS, 1):
            c = ws.cell(row=ri, column=ci, value=lead.get(key, ""))
            c.font = df; c.fill = fill; c.border = tb
            c.alignment = Alignment(vertical="center", wrap_text=(key == "message"))
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A3"

    # Stats sheet
    ws2 = wb.create_sheet("Résumé")
    rows = [
        ("Projet",            project),
        ("Total leads",       len(leads)),
        ("Export généré le",  datetime.now().strftime("%d/%m/%Y à %H:%M")),
        ("Site cible",        cfg.get("target_url", "")),
        ("Backend",           cfg.get("backend_type", "none")),
    ]
    for i, (k, v) in enumerate(rows, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True, name="Calibri", size=10)
        ws2.cell(row=i, column=2, value=v).font = Font(name="Calibri", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 36

    wb.save(path)
    return path
